"""Excel output writer.

Produces the final ``.xlsx`` using :mod:`openpyxl`.  The primary **Results**
sheet has one row per athlete — identified by the composite key
``(NAME, ID, COLLEGE)`` — showing GENDER and the summed SCORE, sorted highest
first.  Two supporting sheets are added:

* **Details**  — every individual performance and its score, grouped under the
  athlete it belongs to, so officials can see how each total was built.
* **Rejected** — every skipped input row with the reason it was excluded.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from athletics_scoring.models import (
    AGGREGATE_COLUMNS,
    OUTPUT_COLUMNS,
    AthleteAggregate,
    ScoringReport,
)

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_DETAIL_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_REJECT_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")


class ExcelWriter:
    """Writes scoring results to a formatted ``.xlsx`` workbook."""

    def write(
        self,
        aggregates: list[AthleteAggregate],
        report: ScoringReport,
        path: Path | str,
        include_details: bool = True,
        include_rejects: bool = True,
    ) -> Path:
        """Write results to *path*.

        Args:
            aggregates: One summed row per athlete, already sorted by descending
                score.
            report: The full scoring report (used for the Details and Rejected
                sheets).
            path: Destination ``.xlsx`` path.
            include_details: When ``True``, add a per-performance Details sheet.
            include_rejects: When ``True`` and rejects exist, add a Rejected
                sheet.

        Returns:
            The resolved output path.
        """
        workbook = Workbook()
        results_sheet = workbook.active
        results_sheet.title = "Results"
        self._write_results(results_sheet, aggregates)

        if include_details:
            detail_sheet = workbook.create_sheet("Details")
            self._write_details(detail_sheet, aggregates)

        if include_rejects and report.rejected:
            reject_sheet = workbook.create_sheet("Rejected")
            self._write_rejects(reject_sheet, report)

        out_path = Path(path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_path)
        return out_path

    # -- Sheets --------------------------------------------------------------
    def _write_results(
        self, sheet: Worksheet, aggregates: list[AthleteAggregate]
    ) -> None:
        """Write the aggregated, one-row-per-athlete Results sheet."""
        headers = ("RANK",) + AGGREGATE_COLUMNS
        sheet.append(list(headers))
        for rank, aggregate in enumerate(aggregates, start=1):
            row = aggregate.as_output_row()
            sheet.append([rank] + [row[col] for col in AGGREGATE_COLUMNS])
        self._style_header(sheet, len(headers), _HEADER_FILL)
        self._autofit(sheet, headers)
        sheet.freeze_panes = "A2"

    def _write_details(
        self, sheet: Worksheet, aggregates: list[AthleteAggregate]
    ) -> None:
        """Write the per-performance breakdown grouped by athlete.

        Athletes appear in the same order as the Results sheet; within each
        athlete the events are ordered from highest to lowest score.
        """
        headers = ("RANK",) + OUTPUT_COLUMNS
        sheet.append(list(headers))
        for rank, aggregate in enumerate(aggregates, start=1):
            for result in aggregate.results:
                row = result.as_output_row()
                sheet.append([rank] + [row[col] for col in OUTPUT_COLUMNS])
        self._style_header(sheet, len(headers), _DETAIL_FILL)
        self._autofit(sheet, headers)
        sheet.freeze_panes = "A2"

    def _write_rejects(self, sheet: Worksheet, report: ScoringReport) -> None:
        """Write the rejected-records sheet."""
        headers = (
            "ROW",
            "NAME",
            "ID",
            "COLLEGE",
            "GENDER",
            "EVENT NAME",
            "PERFORMANCE TYPE",
            "RESULT",
            "REASON",
        )
        sheet.append(list(headers))
        for reject in report.rejected:
            row = reject.as_output_row()
            sheet.append([row[col] for col in headers])
        self._style_header(sheet, len(headers), _REJECT_FILL)
        self._autofit(sheet, headers)
        sheet.freeze_panes = "A2"

    # -- Styling helpers -----------------------------------------------------
    def _style_header(self, sheet: Worksheet, ncols: int, fill: PatternFill) -> None:
        """Apply fill / font / alignment to the header row."""
        for col in range(1, ncols + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = fill
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit(self, sheet: Worksheet, headers: tuple[str, ...]) -> None:
        """Set a reasonable width per column based on its contents."""
        for idx, header in enumerate(headers, start=1):
            letter = get_column_letter(idx)
            max_len = len(str(header))
            for cell in sheet[letter]:
                value = cell.value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            sheet.column_dimensions[letter].width = min(max_len + 2, 40)
