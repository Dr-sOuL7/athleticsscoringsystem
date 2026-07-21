"""Build step: convert the World Athletics workbook into bundled JSON.

Running this module once turns the 2.3 MB, 26-sheet ``.xlsx`` into a single
compact ``scoring_tables.json`` that the application loads at runtime.  This
keeps start-up fast and removes any runtime dependency on the original
spreadsheet.

Output schema
-------------
.. code-block:: json

    {
      "meta":  {"edition": "2025", "source": "...", "generated": "..."},
      "events": {
        "M": {
          "100m": {"type": "TIME",  "perf": [9.46, ...], "pts": [1400, ...]},
          "LJ":   {"type": "DISTANCE", "perf": [...], "pts": [...]}
        },
        "W": { ... }
      }
    }

For every event the ``perf`` array is sorted **ascending** and ``pts`` is the
parallel array of point values, so the runtime can binary-search it directly.

Run with::

    python -m athletics_scoring.build_tables            # default paths
    python -m athletics_scoring.build_tables SRC.xlsx OUT.json
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import sys
from pathlib import Path

import openpyxl

from athletics_scoring.events import PerformanceType
from athletics_scoring.performance import is_empty_token, parse_performance

# Columns on the "Field Events" sheets that are combined-event *point* scores
# rather than metric distances.
_COMBINED_EVENT_CODES: frozenset[str] = frozenset(
    {"Hept. sh", "Dec.", "Pent. Sh", "Heptathlon"}
)

# Sheet-name substrings that indicate a distance-measured (field) sheet.  Every
# other sheet is time-measured.
_FIELD_SHEET_MARKER = "Field Events"


def _gender_key(sheet_name: str) -> str | None:
    """Return ``"M"``/``"W"`` for a sheet, or ``None`` if undetermined."""
    if sheet_name.startswith("Men"):
        return "M"
    if sheet_name.startswith("Women"):
        return "W"
    return None


def _column_type(sheet_name: str, event_code: str) -> PerformanceType:
    """Classify a single (sheet, event) column into a performance type."""
    if _FIELD_SHEET_MARKER in sheet_name:
        if event_code in _COMBINED_EVENT_CODES:
            return PerformanceType.POINTS
        return PerformanceType.DISTANCE
    return PerformanceType.TIME


def _clean_code(raw_header: str) -> str:
    """Normalise a header cell into a canonical event code.

    Trailing/leading whitespace is stripped and internal runs of whitespace are
    collapsed to a single space (``"200m sh "`` -> ``"200m sh"``).
    """
    return " ".join(str(raw_header).split())


def build(source_xlsx: Path, output_json: Path) -> dict:
    """Parse *source_xlsx* and write the compact table JSON to *output_json*.

    Returns:
        The in-memory dictionary that was serialised (useful for tests).
    """
    workbook = openpyxl.load_workbook(source_xlsx, read_only=True, data_only=True)

    # events[gender][code] -> {"type", "pairs": list[(perf_float, points_int)]}
    events: dict[str, dict[str, dict]] = {"M": {}, "W": {}}

    for sheet_name in workbook.sheetnames:
        gender = _gender_key(sheet_name)
        if gender is None:
            continue
        sheet = workbook[sheet_name]

        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        # Column 0 is "Points"; map remaining columns to cleaned event codes.
        col_codes: dict[int, str] = {}
        for idx, cell in enumerate(header):
            if idx == 0 or cell is None:
                continue
            col_codes[idx] = _clean_code(cell)

        # Prepare a per-event accumulator for this sheet.
        buckets: dict[str, list[tuple[float, int]]] = {
            code: [] for code in col_codes.values()
        }
        types: dict[str, PerformanceType] = {
            code: _column_type(sheet_name, code) for code in col_codes.values()
        }

        for row in rows:
            points_cell = row[0]
            if points_cell is None:
                continue
            try:
                points = int(str(points_cell).strip())
            except (TypeError, ValueError):
                continue
            for idx, code in col_codes.items():
                value = row[idx] if idx < len(row) else None
                if is_empty_token(value):
                    continue
                try:
                    perf = parse_performance(value)
                except Exception:
                    # A malformed table cell is skipped rather than aborting the
                    # whole build; such cells are extremely rare.
                    continue
                buckets[code].append((perf, points))

        # Merge this sheet's data into the global structure.
        for code, pairs in buckets.items():
            if not pairs:
                continue
            record = events[gender].setdefault(
                code, {"type": types[code].value, "pairs": []}
            )
            record["pairs"].extend(pairs)

    workbook.close()

    # Finalise: sort by performance ascending and split into parallel arrays.
    payload_events: dict[str, dict[str, dict]] = {"M": {}, "W": {}}
    for gender, code_map in events.items():
        for code, record in code_map.items():
            pairs = sorted(record["pairs"], key=lambda pair: pair[0])
            # De-duplicate identical performances, keeping the LOWER score so
            # that a performance sitting exactly on a shared value is never
            # over-credited (consistent with the "use the lower score" rule).
            perf_list: list[float] = []
            pts_list: list[int] = []
            last_perf: float | None = None
            for perf, pts in pairs:
                if perf == last_perf:
                    pts_list[-1] = min(pts_list[-1], pts)
                    continue
                perf_list.append(perf)
                pts_list.append(pts)
                last_perf = perf
            payload_events[gender][code] = {
                "type": record["type"],
                "perf": perf_list,
                "pts": pts_list,
            }

    payload = {
        "meta": {
            "edition": "2025",
            "source": source_xlsx.name,
            "generated": _dt.datetime.now(_dt.timezone.utc).isoformat(),
            "event_count_men": len(payload_events["M"]),
            "event_count_women": len(payload_events["W"]),
        },
        "events": payload_events,
    }

    output_json.parent.mkdir(parents=True, exist_ok=True)
    with output_json.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, separators=(",", ":"), ensure_ascii=False)

    return payload


def _parse_args(argv: list[str]) -> argparse.Namespace:
    root = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description="Build bundled scoring-table JSON.")
    parser.add_argument(
        "source",
        nargs="?",
        default=str(root / "data" / "source" / "WA_Scoring_Tables_2025.xlsx"),
        help="Path to the World Athletics scoring workbook (.xlsx).",
    )
    parser.add_argument(
        "output",
        nargs="?",
        default=str(root / "data" / "scoring_tables.json"),
        help="Destination path for the generated JSON.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """CLI entry point for the build step."""
    args = _parse_args(argv if argv is not None else sys.argv[1:])
    source = Path(args.source)
    output = Path(args.output)
    if not source.exists():
        print(f"ERROR: source workbook not found: {source}", file=sys.stderr)
        return 1
    payload = build(source, output)
    meta = payload["meta"]
    print(
        f"Built {output} "
        f"({meta['event_count_men']} men / {meta['event_count_women']} women events)"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
